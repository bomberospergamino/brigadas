#!/usr/bin/env python3
"""Prepara el workbook del modulo Brigadas.

Uso:
  python scripts/setup_brigadas_workbook.py archivo.xlsx

Si no se indica archivo, crea/actualiza brigadas_template.xlsx en la carpeta actual.
El script es idempotente: no borra datos, no duplica hojas ni filas base.
"""

from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


BRIGADAS = [
    ["rescate_acuatico", "Rescate acuatico", "rescate_acuatico", "rescateacuatico_logo.jpeg", "#0077B6", "SI", 1, 1, "", "", ""],
    ["buceo", "Buceo", "buceo", "buceo_logo.jpeg", "#003566", "SI", 2, 1, "", "", ""],
    ["k9", "K9", "k9", "k9_logo.jpeg", "#D00000", "SI", 3, 1, "", "", ""],
    ["mat_pel", "Mat Pel", "mat_pel", "matpel_logo.jpeg", "#70E000", "SI", 4, 1, "", "", ""],
    ["altura", "Altura", "altura", "altura_logo.jpeg", "#B5179E", "SI", 5, 1, "", "", ""],
    ["socorrismo", "Socorrismo", "socorrismo", "socorrismo_logo.jpeg", "#E85D04", "SI", 6, 1, "", "", ""],
    ["brec", "BREC", "brec", "brec_logo.jpeg", "#6C584C", "SI", 7, 1, "", "", ""],
]

SHEETS = {
    "CONFIG_BRIGADAS": [
        "id_brigada", "nombre_brigada", "columna_personal", "logo_file", "color", "activa", "orden",
        "frecuencia_minima_mensual", "responsable", "bibliografia_url", "observaciones",
    ],
    "PERSONAL": [
        "id_persona", "apellido", "nombre", "bombero", "legajo", "jerarquia", "activo",
        "rescate_acuatico", "buceo", "k9", "mat_pel", "altura", "socorrismo", "brec", "cantidad_brigadas",
    ],
    "ENCUENTROS": [
        "id_encuentro", "fecha", "mes_periodo", "id_brigada", "brigada", "tipo_encuentro", "tema",
        "responsable", "lugar", "hora_inicio", "hora_fin", "estado", "observaciones", "id_asistencia",
    ],
    "ASISTENCIAS": [
        "id_asistencia", "id_encuentro", "fecha", "mes_periodo", "id_brigada", "brigada", "tipo_actividad",
        "responsable", "observaciones_generales", "creado_por", "timestamp", "pdf_url", "estado_registro",
    ],
    "DETALLE_ASISTENCIAS": [
        "id_asistencia", "legajo", "apellido", "nombre", "bombero", "jerarquia", "estado_asistencia",
        "observacion_individual",
    ],
    "EQUIPAMIENTO": [
        "brigada", "ubicacion", "elemento", "unidades", "cantidad_ok_no", "estado_bueno_malo", "observaciones",
    ],
    "CHECK_EQUIPAMIENTO": [
        "id_check", "fecha", "mes_periodo", "brigada", "ubicacion", "elemento", "unidades", "cantidad_ok_no",
        "estado_bueno_malo", "observaciones", "responsable", "pdf_url",
    ],
    "NECESIDADES": [
        "id_necesidad", "fecha_alta", "id_brigada", "brigada", "tipo_necesidad", "elemento", "descripcion",
        "cantidad", "prioridad", "estado", "presupuesto_unitario", "presupuesto_total", "proveedor",
        "link_referencia", "justificacion_operativa", "origen", "id_check_origen", "responsable",
        "fecha_objetivo", "observaciones",
    ],
    "PARAMETROS": ["tipo_parametro", "valor", "descripcion", "activo", "orden"],
}

UNUSED_SHEETS = ["DETALLE_CHECK_EQUIPAMIENTO", "HISTORIAL_ACCIONES", "ADMIN_RESUMEN"]

LEGACY_HEADERS = {
    "apellido": "apellido",
    "nombre": "nombre",
    "bombero": "bombero",
    "legajo": "legajo",
    "jerarquia": "jerarquia",
    "rescate acuatico": "rescate_acuatico",
    "buceo": "buceo",
    "k9": "k9",
    "mat pel": "mat_pel",
    "altura": "altura",
    "socorrismo": "socorrismo",
    "brec": "brec",
}

PARAMETROS = [
    ["estado_encuentro", "Programado", "Encuentro planificado", "SI", 1],
    ["estado_encuentro", "Realizado", "Encuentro realizado", "SI", 2],
    ["estado_encuentro", "Suspendido", "Encuentro suspendido", "SI", 3],
    ["estado_encuentro", "Reprogramado", "Encuentro reprogramado", "SI", 4],
    ["estado_encuentro", "Pendiente", "Pendiente de definir", "SI", 5],
    ["estado_asistencia", "P", "Presente", "SI", 1],
    ["estado_asistencia", "A", "Ausente", "SI", 2],
    ["estado_asistencia", "Just.", "Justificado", "SI", 3],
    ["estado_equipamiento", "OK", "Correcto", "SI", 1],
    ["estado_equipamiento", "Faltante", "Elemento faltante", "SI", 2],
    ["estado_equipamiento", "Averiado", "Elemento averiado", "SI", 3],
    ["estado_equipamiento", "Vencido", "Elemento vencido", "SI", 4],
    ["estado_equipamiento", "No aplica", "No aplica", "SI", 5],
    ["prioridad", "Alta", "Prioridad alta", "SI", 1],
    ["prioridad", "Media", "Prioridad media", "SI", 2],
    ["prioridad", "Baja", "Prioridad baja", "SI", 3],
    ["estado_necesidad", "Pendiente", "Necesidad pendiente", "SI", 1],
    ["estado_necesidad", "Cotizado", "Ya tiene cotizacion", "SI", 2],
    ["estado_necesidad", "Solicitado", "Pedido solicitado", "SI", 3],
    ["estado_necesidad", "Comprado", "Comprado", "SI", 4],
    ["estado_necesidad", "Recibido", "Recibido", "SI", 5],
    ["estado_necesidad", "Cancelado", "Cancelado", "SI", 6],
    ["resultado_check", "OK", "Todo correcto", "SI", 1],
    ["resultado_check", "Observado", "Hay observaciones", "SI", 2],
    ["resultado_check", "Critico", "Situacion critica", "SI", 3],
    ["activo", "SI", "Registro activo", "SI", 1],
    ["activo", "NO", "Registro inactivo", "SI", 2],
]


def normalize(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return " ".join(text.replace("_", " ").split())


def header_map(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1] if cell.value}


def ensure_sheet(wb, name: str, headers: list[str]):
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        for idx, header in enumerate(headers, 1):
            ws.cell(1, idx, header)
        return ws

    existing = header_map(ws)
    normalized_existing = {normalize(k): (k, v) for k, v in existing.items()}

    if name == "PERSONAL":
        for legacy_key, new_header in LEGACY_HEADERS.items():
            if legacy_key in normalized_existing and new_header not in existing:
                old_header, col = normalized_existing[legacy_key]
                ws.cell(1, col, new_header)
                existing.pop(old_header, None)
                existing[new_header] = col

    existing = header_map(ws)
    next_col = ws.max_column + 1
    for header in headers:
        if header not in existing:
            ws.cell(1, next_col, header)
            next_col += 1
    return ws


def compact_sheet_to_headers(ws, headers: list[str]):
    existing = [cell.value for cell in ws[1]]
    if not any(existing):
        for idx, header in enumerate(headers, 1):
            ws.cell(1, idx, header)
        return
    normalized = [normalize(value) for value in existing]

    def find_index(target: str) -> int | None:
        exact = normalize(target)
        if exact in normalized:
            return normalized.index(exact) + 1
        if target == "ubicacion":
            for index, header in enumerate(normalized, 1):
                if "ubicacion" in header:
                    return index
        if target == "cantidad_ok_no":
            for index, header in enumerate(normalized, 1):
                if "cantidad" in header:
                    return index
        if target == "estado_bueno_malo":
            for index, header in enumerate(normalized, 1):
                if "estado" in header:
                    return index
        return None

    rows = []
    for row in range(2, ws.max_row + 1):
        compact = []
        for header in headers:
            col_idx = find_index(header)
            compact.append(ws.cell(row, col_idx).value if col_idx else "")
        if any(value not in (None, "") for value in compact):
            rows.append(compact)

    ws.delete_cols(1, ws.max_column)
    for idx, header in enumerate(headers, 1):
        ws.cell(1, idx, header)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row_idx, col_idx, value)


def upsert_rows(ws, rows: list[list[object]], key_columns: int):
    existing_keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = tuple(row[:key_columns])
        if any(key):
            existing_keys.add(key)
    for row in rows:
        key = tuple(row[:key_columns])
        if key not in existing_keys:
            ws.append(row)


def list_validation(values: list[str]) -> DataValidation:
    quoted = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    dv.error = "Valor no permitido"
    dv.errorTitle = "Validacion"
    return dv


def apply_format(ws):
    header_fill = PatternFill("solid", fgColor="7A1113")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D6D6D6")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 4, 14), 32)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def col(ws, header: str) -> str | None:
    mapping = header_map(ws)
    if header not in mapping:
        return None
    return get_column_letter(mapping[header])


def add_validation(ws, header: str, dv: DataValidation, start=2, end=1000):
    letter = col(ws, header)
    if not letter:
        return
    ws.add_data_validation(dv)
    dv.add(f"{letter}{start}:{letter}{end}")


def add_formula(ws, header: str, formula_builder, start=2, end=500):
    letter = col(ws, header)
    if not letter:
        return
    for row in range(start, end + 1):
        if ws[f"{letter}{row}"].value in (None, ""):
            ws[f"{letter}{row}"] = formula_builder(row)


def prepare_workbook(path: Path):
    wb = load_workbook(path) if path.exists() else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and not wb["Sheet"]["A1"].value:
        wb.remove(wb["Sheet"])
    for name in UNUSED_SHEETS:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    for name, headers in SHEETS.items():
        ensure_sheet(wb, name, headers)
    compact_sheet_to_headers(wb["EQUIPAMIENTO"], SHEETS["EQUIPAMIENTO"])
    compact_sheet_to_headers(wb["CHECK_EQUIPAMIENTO"], SHEETS["CHECK_EQUIPAMIENTO"])

    upsert_rows(wb["CONFIG_BRIGADAS"], BRIGADAS, 1)
    upsert_rows(wb["PARAMETROS"], PARAMETROS, 2)

    # Corrige columna_personal si una carga previa vino con el nombre de archivo.
    ws_config = wb["CONFIG_BRIGADAS"]
    config_cols = header_map(ws_config)
    for row in range(2, ws_config.max_row + 1):
        brigade_id = ws_config.cell(row, config_cols["id_brigada"]).value
        expected = next((item[2] for item in BRIGADAS if item[0] == brigade_id), None)
        if expected:
            ws_config.cell(row, config_cols["columna_personal"], expected)

    for row in range(2, wb["PERSONAL"].max_row + 1):
        cols = header_map(wb["PERSONAL"])
        if "id_persona" in cols and not wb["PERSONAL"].cell(row, cols["id_persona"]).value and "legajo" in cols:
            wb["PERSONAL"].cell(row, cols["id_persona"], wb["PERSONAL"].cell(row, cols["legajo"]).value)
        if "activo" in cols and not wb["PERSONAL"].cell(row, cols["activo"]).value:
            wb["PERSONAL"].cell(row, cols["activo"], "SI")

    for ws in wb.worksheets:
        apply_format(ws)

    yes_no = list_validation(["SI", "NO"])
    add_validation(wb["PERSONAL"], "activo", yes_no)
    for brigade in [item[2] for item in BRIGADAS]:
        add_validation(wb["PERSONAL"], brigade, list_validation(["x", "X", "SI"]))

    add_validation(wb["CONFIG_BRIGADAS"], "activa", yes_no)
    add_validation(wb["ENCUENTROS"], "estado", list_validation(["Programado", "Realizado", "Suspendido", "Reprogramado", "Pendiente"]))
    add_validation(wb["ASISTENCIAS"], "estado_registro", list_validation(["Borrador", "Cerrado", "Anulado"]))
    add_validation(wb["DETALLE_ASISTENCIAS"], "estado_asistencia", list_validation(["P", "A", "Just."]))
    add_validation(wb["EQUIPAMIENTO"], "cantidad_ok_no", list_validation(["OK", "NO"]))
    add_validation(wb["EQUIPAMIENTO"], "estado_bueno_malo", list_validation(["BUENO", "MALO"]))
    add_validation(wb["CHECK_EQUIPAMIENTO"], "resultado_global", list_validation(["OK", "Observado", "Critico"]))
    add_validation(wb["CHECK_EQUIPAMIENTO"], "cantidad_ok_no", list_validation(["OK", "NO"]))
    add_validation(wb["CHECK_EQUIPAMIENTO"], "estado_bueno_malo", list_validation(["BUENO", "MALO"]))
    add_validation(wb["NECESIDADES"], "prioridad", list_validation(["Alta", "Media", "Baja"]))
    add_validation(wb["NECESIDADES"], "estado", list_validation(["Pendiente", "Cotizado", "Solicitado", "Comprado", "Recibido", "Cancelado"]))

    for sheet_name in ("ENCUENTROS", "ASISTENCIAS", "CHECK_EQUIPAMIENTO"):
        ws = wb[sheet_name]
        fecha_col = col(ws, "fecha")
        if fecha_col:
            add_formula(ws, "mes_periodo", lambda r, fc=fecha_col: f'=IF({fc}{r}<>"",TEXT({fc}{r},"yyyy-mm"),"")')

    personal = wb["PERSONAL"]
    brigade_cols = [col(personal, item[2]) for item in BRIGADAS]
    brigade_cols = [c for c in brigade_cols if c]
    add_formula(
        personal,
        "cantidad_brigadas",
        lambda r: "=" + "+".join([f'IF(OR(LOWER(TRIM({c}{r}))="x",LOWER(TRIM({c}{r}))="si"),1,0)' for c in brigade_cols]),
    )

    nec = wb["NECESIDADES"]
    cantidad_col = col(nec, "cantidad")
    unitario_col = col(nec, "presupuesto_unitario")
    if cantidad_col and unitario_col:
        add_formula(nec, "presupuesto_total", lambda r: f'=IF(AND({cantidad_col}{r}<>"",{unitario_col}{r}<>""),{cantidad_col}{r}*{unitario_col}{r},"")')

    wb.save(path)
    return path


def main():
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("brigadas_template.xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    print(f"Preparando workbook: {target.resolve()}")
    prepare_workbook(target)
    print("Listo. Hojas y estructura del modulo Brigadas preparadas.")


if __name__ == "__main__":
    main()
